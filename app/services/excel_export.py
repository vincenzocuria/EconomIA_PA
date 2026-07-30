from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from app.models.buono import BuonoEconomale
from app.models.cassetto import SaldoAnnuale
from app.models.movimento import Movimento
from app.services.cassa import effetto_su_cassa, effetto_su_conto, movimenti_contabili
from app.services.movimento_display import testo_filiale_per_movimento
from app.services.numero_display import formato_numero_sezionale


def export_movimenti_excel(anno: int) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Movimenti {anno}"
    headers = [
        "N.",
        "Sezionale",
        "Data",
        "Ora",
        "Tipo",
        "Importo",
        "Causale",
        "Beneficiario",
        "CF/P.IVA",
        "Doc. fiscale",
        "Data doc.",
        "Pagamento",
        "Filiale banca",
        "Rif. ricevuta",
        "Capitolo",
        "Stato",
        "Trimestre",
        "Buono ID",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
    rows = Movimento.query.filter_by(anno=anno).order_by(Movimento.numero_progressivo).all()
    r = 2
    for m in rows:
        ws.cell(row=r, column=1, value=formato_numero_sezionale(m))
        ws.cell(row=r, column=2, value=m.sezionale.codice if m.sezionale else "")
        ws.cell(row=r, column=3, value=m.data_movimento)
        ws.cell(row=r, column=4, value=m.ora_movimento.strftime("%H:%M") if m.ora_movimento else "")
        ws.cell(row=r, column=5, value=m.tipo.value if hasattr(m.tipo, "value") else str(m.tipo))
        ws.cell(row=r, column=6, value=float(m.importo))
        ws.cell(row=r, column=7, value=m.causale)
        ws.cell(row=r, column=8, value=m.beneficiario_fornitore)
        ws.cell(row=r, column=9, value=m.cf_piva)
        ws.cell(row=r, column=10, value=m.num_documento_fiscale)
        ws.cell(row=r, column=11, value=m.data_documento_fiscale)
        ws.cell(row=r, column=12, value=m.modalita_pagamento)
        ws.cell(row=r, column=13, value=testo_filiale_per_movimento(m))
        ws.cell(row=r, column=14, value=getattr(m, "rif_ricevuta", "") or "")
        ws.cell(row=r, column=15, value=m.capitolo_riferimento)
        ws.cell(row=r, column=16, value=m.stato.value if hasattr(m.stato, "value") else str(m.stato))
        ws.cell(row=r, column=17, value=m.trimestre)
        ws.cell(row=r, column=18, value=m.buono_id)
        r += 1
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def export_buoni_excel(anno: int) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Buoni {anno}"
    headers = ["N.", "Sezionale", "Data", "Stato", "Richiedente", "Ufficio", "Autorizzato", "Speso", "Beneficiario"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True)
    rows = BuonoEconomale.query.filter_by(anno=anno).order_by(BuonoEconomale.numero_progressivo).all()
    r = 2
    for b in rows:
        ws.cell(row=r, column=1, value=formato_numero_sezionale(b))
        ws.cell(row=r, column=2, value=b.sezionale.codice if b.sezionale else "")
        ws.cell(row=r, column=3, value=b.data_buono)
        ws.cell(row=r, column=4, value=b.stato.value if hasattr(b.stato, "value") else str(b.stato))
        ws.cell(row=r, column=5, value=b.richiedente)
        ws.cell(row=r, column=6, value=b.ufficio_richiedente)
        ws.cell(row=r, column=7, value=float(b.importo_autorizzato))
        ws.cell(row=r, column=8, value=float(b.importo_speso))
        ws.cell(row=r, column=9, value=b.beneficiario)
        r += 1
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def nome_file_export(prefix: str, anno: int) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return f"{prefix}_{anno}_{ts}.xlsx"


def export_riepilogo_annuale_excel(anno: int) -> BytesIO:
    from decimal import Decimal

    wb = Workbook()
    ws = wb.active
    ws.title = f"Riepilogo {anno}"
    saldo_row = SaldoAnnuale.query.get(anno)
    ini_cassa = Decimal(str(saldo_row.saldo_iniziale)) if saldo_row else Decimal("0")
    ini_conto = (
        Decimal(str(getattr(saldo_row, "saldo_conto_iniziale", 0) or 0)) if saldo_row else Decimal("0")
    )
    ws.append(["Riepilogo annuale cassa / conto economale", anno])
    ws.append(["Saldo iniziale cassa (contanti)", float(ini_cassa)])
    ws.append(["Saldo iniziale conto (banca)", float(ini_conto)])
    per_t_cassa = {1: Decimal("0"), 2: Decimal("0"), 3: Decimal("0"), 4: Decimal("0")}
    per_t_conto = {1: Decimal("0"), 2: Decimal("0"), 3: Decimal("0"), 4: Decimal("0")}
    for m in movimenti_contabili(anno):
        per_t_cassa[m.trimestre] = per_t_cassa[m.trimestre] + effetto_su_cassa(m)
        per_t_conto[m.trimestre] = per_t_conto[m.trimestre] + effetto_su_conto(m)
    tot_cassa = sum(per_t_cassa.values(), start=Decimal("0"))
    tot_conto = sum(per_t_conto.values(), start=Decimal("0"))
    ws.append([])
    ws.append(["Trimestre", "Variazione cassa", "Variazione conto"])
    for t in range(1, 5):
        ws.append([f"T{t}", float(per_t_cassa[t]), float(per_t_conto[t])])
    ws.append(["Totale movimenti", float(tot_cassa), float(tot_conto)])
    ws.append(["Saldo finale cassa stimato", float(ini_cassa + tot_cassa)])
    ws.append(["Saldo finale conto stimato", float(ini_conto + tot_conto)])
    ws.append([])
    ws.append(["Nota", "Calcolo da registro locale; verificare con contabilità ufficiale e estratto conto."])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
